import sys
from pathlib import Path
import openpyxl

def inspect(path, max_sheets=50, sample_rows=5):
    print(f"\n{'='*100}\nFILE: {path.name}  ({path.stat().st_size:,} bytes)\n{'='*100}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening: {e}")
        return
    print(f"  Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    for sn in wb.sheetnames[:max_sheets]:
        ws = wb[sn]
        print(f"\n  --- Sheet: {sn!r}  rows~{ws.max_row} cols~{ws.max_column}")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= sample_rows + 5:
                break
        for i, row in enumerate(rows):
            vals = ["" if v is None else str(v) for v in row]
            line = " | ".join(vals)
            if len(line) > 240:
                line = line[:240] + "…"
            print(f"    [{i}] {line}")
    wb.close()

if __name__ == "__main__":
    targets = sys.argv[1:]
    for t in targets:
        inspect(Path(t))
