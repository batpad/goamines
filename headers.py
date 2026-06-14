import sys
from pathlib import Path
import openpyxl

def headers(path):
    print(f"\n### {path.name}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr = None
        for row in ws.iter_rows(values_only=True):
            hdr = row
            break
        cols = [("" if v is None else str(v)) for v in (hdr or ())]
        # count data rows cheaply
        print(f"  [{sn}] rows~{ws.max_row}")
        for i, c in enumerate(cols):
            print(f"      {i:2d}. {c}")
    wb.close()

for t in sys.argv[1:]:
    headers(Path(t))
