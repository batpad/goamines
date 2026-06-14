"""
Ingest Goa ore-transport data (data1 permits + data2 trips/closing-stock) into SQLite.

Run:  uv run python ingest.py
Output: goamines.db
"""
import re
import sqlite3
from pathlib import Path
from datetime import datetime, timedelta, date
import openpyxl

ROOT = Path(__file__).parent
D1 = ROOT / "data1"
D2 = ROOT / "data2"
DB = ROOT / "goamines.db"

EXCEL_EPOCH = datetime(1899, 12, 30)

# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
def norm_header(h):
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = s.replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.replace(" : ", ": ")
    return s.strip()

# canonical permit field -> set of normalized header variants
PERMIT_SYNONYMS = {
    "permit_no":        {"permit no", "permit number"},
    "permit_type":      {"permit type"},
    "issue_date":       {"issue date", "permit issue date"},
    "validity_date":    {"validity date"},
    "application_date": {"application date"},
    "status":           {"permit status"},
    "financial_year":   {"financial year"},
    "org_code":         {"organization code", "code"},
    "org_name":         {"organization name", "organisation/trader name", "party name"},
    "mineral_type":     {"mineral type", "mineral"},
    "ore_type":         {"ore type"},
    "ore_category":     {"ore category"},
    "grade_slab":       {"grade slab"},
    "exact_grade":      {"exact grade"},
    "permit_qty":       {"permit quantity", "mineral quantity", "permit proposed quantity"},
    "used_qty":         {"used quantity"},
    "balance_qty":      {"permit balance"},
    "source_location":  {"source location"},
    "dest_location":    {"destination location", "destination"},
    "transport_mode":   {"transportation mode", "transportation",
                         "transportation mode: source to destination",
                         "transportation: source to destination",
                         "transportation mode : source to destination",
                         "transportation mode: source to destination"},
    "state":            {"state"},
    "country":          {"country"},
    "district":         {"district"},
    "tehsil":           {"tehsil"},
    "vessel_name":      {"vessel name", "vessel", "vesselname"},
    "buyer_name":       {"buyer name", "buyer", "buyer party name"},
    "trader_name":      {"buying org/trader name"},
    "trader_code":      {"buying org/trader code"},
    "export_sale_permit_no": {"export sale permit no", "export permit number"},
    "export_permit_date":    {"export permit date"},
    "challan_no":       {"challan no"},
    "imp_exp_type":     {"import/export type", "import type", "export type"},
    "imp_exp_address":  {"import/export address", "buyer address"},
    "imp_exp_party":    {"import/export seller/buyer name"},
}
# build reverse lookup: normalized header -> canonical field
HDR2FIELD = {}
for field, variants in PERMIT_SYNONYMS.items():
    for v in variants:
        HDR2FIELD[v] = field

PERMIT_FIELDS = list(PERMIT_SYNONYMS.keys())

def to_iso(v):
    """Normalize a date-ish cell to ISO 'YYYY-MM-DD' (or None)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        # Excel serial date (Megasoft Captive/eAuction store dates as serials)
        if 30000 <= v <= 60000:
            return (EXCEL_EPOCH + timedelta(days=float(v))).date().isoformat()
        return None
    s = str(v).strip()
    if not s or s.startswith("0000-00-00"):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).date().isoformat()
        except ValueError:
            continue
    return None

def to_iso_dt(v):
    """Normalize a datetime cell to ISO 'YYYY-MM-DD HH:MM:SS' (trips)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(microsecond=0).isoformat(sep=" ")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(microsecond=0).isoformat(sep=" ")
        except ValueError:
            continue
    return s

def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None

# stream / era from data1 filename
def parse_d1_meta(fname):
    low = fname.lower()
    if low.startswith("megasoft"):
        era = "megasoft"
    elif low.startswith("old"):
        era = "old_bhumija"
    elif low.startswith("new"):
        era = "new_bhumija"
    else:
        era = "unknown"
    if "captive" in low:
        stream = "imported"   # Captive == Imported stock stream
    elif "royalty" in low:
        stream = "royalty"
    elif "eauction" in low:
        stream = "eauction"
    else:
        stream = "unknown"
    return era, stream

# ----------------------------------------------------------------------------
# schema
# ----------------------------------------------------------------------------
SCHEMA = """
DROP TABLE IF EXISTS permits;
DROP TABLE IF EXISTS trips;
DROP TABLE IF EXISTS closing_stock;
DROP TABLE IF EXISTS locations;
DROP TABLE IF EXISTS location_aliases;

CREATE TABLE permits (
  id INTEGER PRIMARY KEY,
  permit_no TEXT,
  software_era TEXT,
  ore_stream TEXT,
  category TEXT,                 -- transit / export / sale / import
  is_superseded INTEGER DEFAULT 0,
  permit_type TEXT, issue_date TEXT, validity_date TEXT, application_date TEXT,
  status TEXT, financial_year TEXT,
  org_code TEXT, org_name TEXT,
  mineral_type TEXT, ore_type TEXT, ore_category TEXT,
  grade_slab TEXT, exact_grade TEXT,
  permit_qty REAL, used_qty REAL, balance_qty REAL,
  source_location_raw TEXT, dest_location_raw TEXT, transport_mode TEXT,
  state TEXT, country TEXT, district TEXT, tehsil TEXT,
  vessel_name TEXT, buyer_name TEXT, trader_name TEXT, trader_code TEXT,
  export_sale_permit_no TEXT, export_permit_date TEXT, challan_no TEXT,
  imp_exp_type TEXT, imp_exp_address TEXT, imp_exp_party TEXT,
  source_file TEXT, source_sheet TEXT
);

CREATE TABLE trips (
  id INTEGER PRIMARY KEY,
  permit_no TEXT,
  ore_stream TEXT,
  vehicle_or_barge TEXT,
  source_location_raw TEXT, dest_location_raw TEXT,
  start_dt TEXT, end_dt TEXT,
  qty_source REAL, qty_dest REAL, qty_delta REAL,
  source_file TEXT, source_sheet TEXT
);

CREATE TABLE closing_stock (
  id INTEGER PRIMARY KEY,
  as_of_date TEXT,
  ore_stream TEXT,
  location_raw TEXT,
  balance_mt REAL,
  stock_type TEXT,
  source_file TEXT
);

CREATE TABLE locations (
  canonical_name TEXT PRIMARY KEY,
  location_type TEXT,
  state TEXT, country TEXT,
  lat REAL, lon REAL,
  geocode_source TEXT,          -- manual / nominatim / NULL
  geocode_note TEXT,            -- OSM display_name or manual note (auditable)
  notes TEXT
);

CREATE TABLE location_aliases (
  raw_name TEXT PRIMARY KEY,
  canonical_name TEXT REFERENCES locations(canonical_name)
);

-- ---- exploration views (lazy; populated once tables are loaded) ----

-- each trip with its source/dest mapped to canonical locations
CREATE VIEW v_trip_canon AS
SELECT t.*, sa.canonical_name AS src_canon, da.canonical_name AS dst_canon
FROM trips t
LEFT JOIN location_aliases sa ON sa.raw_name = t.source_location_raw
LEFT JOIN location_aliases da ON da.raw_name = t.dest_location_raw;

-- source -> destination route aggregates (canonical), by stream
CREATE VIEW v_routes AS
SELECT src_canon AS source, dst_canon AS destination, ore_stream,
       COUNT(*) AS trips,
       ROUND(SUM(qty_dest),1) AS tonnage,
       ROUND(SUM(qty_source),1) AS tonnage_src,
       ROUND(SUM(qty_dest) - SUM(qty_source),1) AS delta,
       MIN(start_dt) AS first_trip, MAX(start_dt) AS last_trip
FROM v_trip_canon
WHERE src_canon IS NOT NULL AND dst_canon IS NOT NULL
GROUP BY 1,2,3;

-- per-location inflow / outflow / net (canonical), by stream
CREATE VIEW v_location_flows AS
WITH flow AS (
  SELECT dst_canon loc, ore_stream, qty_dest q, 1 is_in FROM v_trip_canon WHERE dst_canon IS NOT NULL
  UNION ALL
  SELECT src_canon loc, ore_stream, qty_source q, 0 is_in FROM v_trip_canon WHERE src_canon IS NOT NULL
)
SELECT loc AS location, ore_stream,
  SUM(is_in) AS trips_in,
  ROUND(SUM(CASE WHEN is_in=1 THEN q END),1) AS tonnage_in,
  SUM(1-is_in) AS trips_out,
  ROUND(SUM(CASE WHEN is_in=0 THEN q END),1) AS tonnage_out,
  ROUND(SUM(CASE WHEN is_in=1 THEN q ELSE -q END),1) AS net_tonnage
FROM flow GROUP BY 1,2;

-- geocoded locations for the map (cluster-map auto-detects latitude/longitude)
CREATE VIEW v_location_map AS
SELECT l.canonical_name AS location, l.location_type, l.state, l.country,
       l.lat AS latitude, l.lon AS longitude, l.geocode_source,
       (SELECT ROUND(SUM(qty_dest),0)   FROM v_trip_canon WHERE dst_canon=l.canonical_name) AS tonnage_in,
       (SELECT ROUND(SUM(qty_source),0) FROM v_trip_canon WHERE src_canon=l.canonical_name) AS tonnage_out
FROM locations l WHERE l.lat IS NOT NULL;

-- permits with their realised trip totals (default snapshot: superseded rows excluded)
CREATE VIEW v_permit_trips AS
SELECT p.permit_no, p.software_era, p.ore_stream, p.category, p.permit_type,
       p.issue_date, p.org_name, p.mineral_type, p.grade_slab, p.permit_qty,
       p.source_location_raw, p.dest_location_raw,
       t.trips, ROUND(t.tonnage,1) AS tonnage_moved
FROM permits p
LEFT JOIN (SELECT permit_no, COUNT(*) trips, SUM(qty_dest) tonnage FROM trips GROUP BY 1) t
       ON t.permit_no = p.permit_no
WHERE p.is_superseded = 0;

-- monthly tonnage time series by stream
CREATE VIEW v_monthly AS
SELECT substr(start_dt,1,7) AS month, ore_stream,
       COUNT(*) AS trips, ROUND(SUM(qty_dest),1) AS tonnage
FROM trips WHERE start_dt IS NOT NULL GROUP BY 1,2;

-- closing stock with canonical location
CREATE VIEW v_closing_stock AS
SELECT cs.as_of_date, cs.ore_stream,
       COALESCE(la.canonical_name, cs.location_raw) AS location, cs.stock_type,
       ROUND(SUM(cs.balance_mt),1) AS balance_mt
FROM closing_stock cs LEFT JOIN location_aliases la ON la.raw_name = cs.location_raw
GROUP BY 1,2,3,4;
"""

# ----------------------------------------------------------------------------
# ingest permits (data1)
# ----------------------------------------------------------------------------
def ingest_permits(con):
    rows_out = []
    for f in sorted(D1.glob("*.xlsx")):
        era, stream = parse_d1_meta(f.name)
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            category = sn.strip().lower()
            it = ws.iter_rows(values_only=True)
            try:
                header = next(it)
            except StopIteration:
                continue
            # map column index -> canonical field
            colmap = {}
            for idx, h in enumerate(header):
                field = HDR2FIELD.get(norm_header(h))
                if field and field not in colmap:   # first match wins
                    colmap[field] = idx
            for r in it:
                if not r or all(c is None for c in r):
                    continue
                rec = {fld: (r[idx] if idx < len(r) else None) for fld, idx in colmap.items()}
                if not rec.get("permit_no"):
                    continue
                rows_out.append((
                    str(rec.get("permit_no")).strip(), era, stream, category, 0,
                    rec.get("permit_type"),
                    to_iso(rec.get("issue_date")), to_iso(rec.get("validity_date")),
                    to_iso(rec.get("application_date")),
                    rec.get("status"), rec.get("financial_year"),
                    rec.get("org_code"), rec.get("org_name"),
                    rec.get("mineral_type"), rec.get("ore_type"), rec.get("ore_category"),
                    rec.get("grade_slab"), rec.get("exact_grade"),
                    to_float(rec.get("permit_qty")), to_float(rec.get("used_qty")),
                    to_float(rec.get("balance_qty")),
                    _clean(rec.get("source_location")), _clean(rec.get("dest_location")),
                    rec.get("transport_mode"),
                    rec.get("state"), rec.get("country"), rec.get("district"), rec.get("tehsil"),
                    rec.get("vessel_name"), rec.get("buyer_name"),
                    rec.get("trader_name"), rec.get("trader_code"),
                    rec.get("export_sale_permit_no"), to_iso(rec.get("export_permit_date")),
                    rec.get("challan_no"),
                    rec.get("imp_exp_type"), rec.get("imp_exp_address"), rec.get("imp_exp_party"),
                    f.name, sn,
                ))
        wb.close()
    con.executemany(f"""INSERT INTO permits
        (permit_no, software_era, ore_stream, category, is_superseded,
         permit_type, issue_date, validity_date, application_date, status, financial_year,
         org_code, org_name, mineral_type, ore_type, ore_category, grade_slab, exact_grade,
         permit_qty, used_qty, balance_qty, source_location_raw, dest_location_raw, transport_mode,
         state, country, district, tehsil, vessel_name, buyer_name, trader_name, trader_code,
         export_sale_permit_no, export_permit_date, challan_no, imp_exp_type, imp_exp_address,
         imp_exp_party, source_file, source_sheet)
        VALUES ({','.join(['?']*40)})""", rows_out)
    con.commit()
    # mark Old-Bhumija rows superseded where the same permit_no exists in New-Bhumija
    con.execute("""
        UPDATE permits SET is_superseded = 1
        WHERE software_era = 'old_bhumija'
          AND permit_no IN (SELECT permit_no FROM permits WHERE software_era = 'new_bhumija')
    """)
    con.commit()
    return len(rows_out)

def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s or None

# ----------------------------------------------------------------------------
# ingest trips (Annexure II)  -- streaming, batched
# ----------------------------------------------------------------------------
TRIP_FILES = [
    (D2 / "Annexure II.xlsx",      {"Royalty": "royalty"}),
    (D2 / "Annexure II (1).xlsx",  {"Imported": "imported", "E-Auction": "eauction"}),
]
def ingest_trips(con):
    total = 0
    for f, sheet_streams in TRIP_FILES:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sn, stream in sheet_streams.items():
            ws = wb[sn]
            batch = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i < 2:        # row 0 title, row 1 header
                    continue
                if not r or r[0] is None:
                    continue
                qs = to_float(r[4]) if len(r) > 4 else None
                qd = to_float(r[7]) if len(r) > 7 else None
                delta = (qd - qs) if (qs is not None and qd is not None) else None
                batch.append((
                    str(r[0]).strip(), stream,
                    _clean(r[1]) if len(r) > 1 else None,
                    _clean(r[2]) if len(r) > 2 else None,
                    _clean(r[5]) if len(r) > 5 else None,
                    to_iso_dt(r[3]) if len(r) > 3 else None,
                    to_iso_dt(r[6]) if len(r) > 6 else None,
                    qs, qd, delta, f.name, sn,
                ))
                if len(batch) >= 5000:
                    _flush_trips(con, batch); total += len(batch); batch = []
            if batch:
                _flush_trips(con, batch); total += len(batch)
        wb.close()
    con.commit()
    return total

def _flush_trips(con, batch):
    con.executemany("""INSERT INTO trips
        (permit_no, ore_stream, vehicle_or_barge, source_location_raw, dest_location_raw,
         start_dt, end_dt, qty_source, qty_dest, qty_delta, source_file, source_sheet)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", batch)

# ----------------------------------------------------------------------------
# ingest closing stock (Annexure III)
# ----------------------------------------------------------------------------
STOCK_FILES = [
    (D2 / "Annexure III-Closing stock of Major Mineral mineral ore as of 31 Mar 2024 (2).xlsx",
     "2024-03-31", {"Imported": "imported", "Eauction": "eauction"}),
    (D2 / "Annexure III.xlsx",
     "2025-03-31", {"Royalty paid": "royalty"}),
    (D2 / "Annexure III (1).xlsx",
     "2025-03-31", {"Imported": "imported", "E-Auction": "eauction"}),
]
def ingest_stock(con):
    rows = []
    for f, as_of, sheet_streams in STOCK_FILES:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sn, stream in sheet_streams.items():
            ws = wb[sn]
            data = list(ws.iter_rows(values_only=True))
            # find header row (contains 'location')
            hr = None
            for idx, r in enumerate(data[:5]):
                if r and any(c and "location" in str(c).lower() for c in r):
                    hr = idx; break
            if hr is None:
                continue
            hdr = [norm_header(c) for c in data[hr]]
            loc_i = next((i for i, h in enumerate(hdr) if "location" in h), 0)
            bal_i = next((i for i, h in enumerate(hdr)
                          if any(k in h for k in ("balance", "quantity", "closing", "stock"))
                          and "type" not in h and "location" not in h), 1)
            type_i = next((i for i, h in enumerate(hdr) if "stock type" in h), None)
            for r in data[hr+1:]:
                if not r or r[loc_i] is None:
                    continue
                loc = _clean(r[loc_i])
                if not loc:
                    continue
                rows.append((
                    as_of, stream, loc,
                    to_float(r[bal_i]) if bal_i < len(r) else None,
                    (_clean(r[type_i]) if (type_i is not None and type_i < len(r)) else None),
                    f.name,
                ))
        wb.close()
    con.executemany("""INSERT INTO closing_stock
        (as_of_date, ore_stream, location_raw, balance_mt, stock_type, source_file)
        VALUES (?,?,?,?,?,?)""", rows)
    con.commit()
    return len(rows)

# ----------------------------------------------------------------------------
# (locations gazetteer is built by locations_build.build(), called from main)
# ----------------------------------------------------------------------------
def main():
    if DB.exists():
        DB.unlink()
    con = sqlite3.connect(DB)
    con.executescript(SCHEMA)
    np = ingest_permits(con)
    print(f"permits:       {np:>7,} rows")
    nt = ingest_trips(con)
    print(f"trips:         {nt:>7,} rows")
    ns = ingest_stock(con)
    print(f"closing_stock: {ns:>7,} rows")
    # indexes
    con.executescript("""
        CREATE INDEX ix_permits_no ON permits(permit_no);
        CREATE INDEX ix_trips_permit ON trips(permit_no);
        CREATE INDEX ix_trips_src ON trips(source_location_raw);
        CREATE INDEX ix_trips_dst ON trips(dest_location_raw);
    """)
    con.commit()
    sup = con.execute("SELECT COUNT(*) FROM permits WHERE is_superseded=1").fetchone()[0]
    print(f"  (marked {sup} old_bhumija permit rows superseded)")
    con.close()
    # curated locations gazetteer (merge variants; geocoding deferred)
    import locations_build
    locations_build.build()
    # precompute per-location balance reconstruction (too slow as live SQL)
    import build_balances
    build_balances.build()
    # materialize the heavy aggregate views into tables (static data -> snappy Datasette)
    con = sqlite3.connect(DB)
    con.executescript("""
        DROP TABLE IF EXISTS routes;
        CREATE TABLE routes AS SELECT * FROM v_routes;
        DROP TABLE IF EXISTS location_flows;
        CREATE TABLE location_flows AS SELECT * FROM v_location_flows;
        DROP TABLE IF EXISTS location_map;
        CREATE TABLE location_map AS SELECT * FROM v_location_map;
        DROP TABLE IF EXISTS monthly;
        CREATE TABLE monthly AS SELECT * FROM v_monthly;
    """)
    con.commit()
    con.close()
    # regenerate the standalone route-arc map
    import build_route_map
    build_route_map.main()
    print(f"\nwrote {DB}")

if __name__ == "__main__":
    main()
