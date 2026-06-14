import re
from pathlib import Path
from collections import defaultdict
import openpyxl

D1 = Path("data1")
seen = defaultdict(list)   # permit_no -> list of (file,sheet)
for f in sorted(D1.glob("*.xlsx")):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        first = True
        for r in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            pn = r[0]
            if pn:
                seen[str(pn).strip()].append((f.name.split("_")[0], sn))
    wb.close()

dupes = {k:v for k,v in seen.items() if len(v) > 1}
print("distinct permits:", len(seen), " with >1 row:", len(dupes))

# categorize the duplication pattern
from collections import Counter
pat = Counter()
for k,v in dupes.items():
    files = tuple(sorted(set(f for f,s in v)))
    sheets = tuple(sorted(set(s for f,s in v)))
    if len(files) > 1:
        pat[("cross-file", files)] += 1
    else:
        pat[("same-file-multi-sheet", sheets)] += 1
for (kind, key), n in pat.most_common(20):
    print(f"  {n:5d}  {kind:22s} {key}")

print("\nexamples:")
for k in list(dupes)[:8]:
    print(" ", k, "->", dupes[k])
