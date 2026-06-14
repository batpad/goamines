import re
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl

D1 = Path("data1")
D2 = Path("data2")

def load(path, header_row=0):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        out[sn] = rows
    wb.close()
    return out

def permit_prefix(pn):
    if not pn:
        return None
    pn = str(pn).strip()
    m = re.match(r'^([A-Z]+)\s*DMG', pn)            # Megasoft/old: "ITP DMG-00743/2020-21"
    if m:
        return m.group(1) + " DMG"
    m = re.match(r'^([A-Z]+)\d', pn)               # Bhumija: "ILT072400024"
    if m:
        return m.group(1) + "#"
    return pn[:6]

# ---- 1. data1 permits: collect all permit numbers + prefixes + date span ----
print("="*80)
print("DATA1 PERMITS — prefixes & permit-number sets")
print("="*80)
d1_permits = set()
d1_prefix = Counter()
for f in sorted(D1.glob("*.xlsx")):
    sheets = load(f)
    for sn, rows in sheets.items():
        if not rows:
            continue
        hdr = [str(c).strip() if c else "" for c in rows[0]]
        # find permit col
        pcol = 0
        for i, h in enumerate(hdr):
            if h.lower().replace("_"," ").replace("number","no") in ("permit no",):
                pcol = i; break
        n = 0
        for r in rows[1:]:
            pn = r[pcol] if pcol < len(r) else None
            if pn:
                d1_permits.add(str(pn).strip())
                d1_prefix[permit_prefix(pn)] += 1
                n += 1
        print(f"  {f.name[:28]:28s} [{sn:8s}] rows={n}")
print("\n  data1 distinct permits:", len(d1_permits))
print("  data1 prefixes:", dict(d1_prefix.most_common()))

# ---- 2. Annexure II trip permits ----
print("\n" + "="*80)
print("ANNEXURE II TRIPS — permit linkage to data1")
print("="*80)
trip_files = [
    (D2/"Annexure II.xlsx", None),
    (D2/"Annexure II (1).xlsx", None),
]
trip_permits = set()
trip_prefix = Counter()
for f, _ in trip_files:
    sheets = load(f)
    for sn, rows in sheets.items():
        # header is row index 1 (row 0 is title)
        cnt = 0
        pcol = 0
        for r in rows[2:]:
            pn = r[pcol] if r else None
            if pn:
                trip_permits.add(str(pn).strip())
                trip_prefix[permit_prefix(pn)] += 1
                cnt += 1
        print(f"  {f.name:22s} [{sn:10s}] trip-rows={cnt}")
print("\n  trip distinct permits:", len(trip_permits))
print("  trip prefixes:", dict(trip_prefix.most_common()))
inter = trip_permits & d1_permits
print(f"\n  trip permits ALSO in data1: {len(inter)} / {len(trip_permits)} "
      f"({100*len(inter)/max(1,len(trip_permits)):.1f}%)")
print("  example trip permits NOT in data1:",
      list(sorted(trip_permits - d1_permits))[:10])

# ---- 3. Locations across everything ----
print("\n" + "="*80)
print("LOCATIONS")
print("="*80)
def collect_locs():
    locs = Counter()
    # data1 source/dest
    for f in sorted(D1.glob("*.xlsx")):
        for sn, rows in load(f).items():
            if not rows: continue
            hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
            for i,h in enumerate(hdr):
                if "source location" in h or h=="destination" or "destination location" in h:
                    for r in rows[1:]:
                        v = r[i] if i<len(r) else None
                        if v: locs[str(v).strip()] += 1
    return locs
locs = collect_locs()
print("  distinct location strings in data1 src/dest:", len(locs))
print("  top 15:", [l for l,_ in locs.most_common(15)])

# closing-stock locations
def stock_locs(path, sheets_rows_hdr):
    out = {}
    for sn, rows in load(path).items():
        # find header row containing 'Location'
        hr = None
        for idx,r in enumerate(rows[:4]):
            if r and any(c and 'location' in str(c).lower() for c in r):
                hr = idx; break
        if hr is None: continue
        names = set()
        for r in rows[hr+1:]:
            if r and r[0]:
                names.add(str(r[0]).strip())
        out[sn] = names
    return out

s2024 = stock_locs(D2/"Annexure III-Closing stock of Major Mineral mineral ore as of 31 Mar 2024 (2).xlsx", None)
s2025a = stock_locs(D2/"Annexure III.xlsx", None)
s2025b = stock_locs(D2/"Annexure III (1).xlsx", None)
print("\n  closing-stock 2024 sheets:", {k:len(v) for k,v in s2024.items()})
print("  closing-stock 2025 sheets:", {**{k:len(v) for k,v in s2025a.items()}, **{k:len(v) for k,v in s2025b.items()}})
all_stock_locs = set()
for d in (s2024,s2025a,s2025b):
    for v in d.values(): all_stock_locs |= v
print("  distinct closing-stock locations:", len(all_stock_locs))
# how many stock locations appear in data1 movement locations?
matched = sum(1 for l in all_stock_locs if l in locs)
print(f"  stock locations exact-matching a data1 movement location: {matched}/{len(all_stock_locs)}")
